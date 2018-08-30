'''
Displays client Wi-Fi statistics in real-time, writes to .xlxs file simultaneously

Useful for testing how a Wi-Fi client roams from AP to AP
Uses standard OS calls to collect connection statistics from NIC driver
Works on Microsoft and Apple platforms
Use keyboard break to stop

Added ICMP check for Microsoft OS.

netpacket.net
'''

__author__ = 'Michal Kowalik'
__credits__= 'Johannes Jobst'
__version__= '1.1'
__status__ = 'Prototype'
__python__ = 'version 3.6.3'

import subprocess, time, re, os
from openpyxl import Workbook

def apple():
    """Outputs data for Apple OS
    """
    wifi = True
    command = 'exec /System/Library/PrivateFrameworks/Apple80211.framework/Versions/Current/Resources/airport -I'
    output = subprocess.check_output(command,shell=True).decode('ascii')
    try:
        ssid = 'SSID:' + re.search(r' SSID.+', output).group(0).split()[-1]
        bssid = 'BSSID:' + re.search(r'BSSID.+', output).group(0).split()[-1]
        channel = 'Ch:' + re.search(r'channel.+', output).group(0).split()[-1]
        txrate = 'TX:' + re.search(r'lastTxRate.+', output).group(0).split()[-1]
        signal = 'Signal:' + re.search(r'CtlRSSI.+', output).group(0).split()[-1]
        noise = 'Noise' + re.search(r'agrCtlNoise.+', output).group(0).split()[-1]
    except AttributeError:
        wifi = False
        print('No RF data')
    if wifi:
        clock = time.asctime().split()[3]
        print(clock, ssid, bssid, channel, txrate, noise, signal)
    time.sleep(0.3)

def microsoft(row):
    """Outputs data for Microsoft OS
    """
    wifi = True
    output = subprocess.check_output('netsh wlan show interfaces').decode('ascii')
    try:
        ping_output = subprocess.check_output(ping_command).decode('ascii')
        ping_output = re.search(r'^Reply.+$', ping_output, re.MULTILINE).group(0)
    except subprocess.CalledProcessError:
        ping_output = 'No ICMP data'
    try:
        ssid = 'SSID:' + re.search(r'SSID.+', output).group(0).split()[-1]
        bssid = 'BSSID:' + re.search(r'BSSID.+', output).group(0).split()[-1]
        channel = 'Ch:' + re.search(r'Channel.+', output).group(0).split()[-1]
        txrate = 'TX:' + re.search(r'Transmit.+', output).group(0).split()[-1]
        rxrate = 'RX:' + re.search(r'Receive.+', output).group(0).split()[-1]
        signal = re.search(r'Signal.+', output).group(0).split()[-1][:-1]
    except AttributeError:
        wifi = False
        print('No RF data')
    if wifi:
        clock = time.asctime().split()[3]
        dbm = signal + '%'
        #dbm = 'dBm:' + str(int(signal) / 2 - 100)
        print(clock, ssid, bssid, dbm, channel, txrate, rxrate)
        print(ping_output)
        xls_sheet.cell(row=row, column=1, value=clock)
        xls_sheet.cell(row=row, column=2, value=ssid)
        xls_sheet.cell(row=row, column=3, value=bssid)
        xls_sheet.cell(row=row, column=4, value=dbm)
        xls_sheet.cell(row=row, column=5, value=channel)
        xls_sheet.cell(row=row, column=6, value=txrate)
        xls_sheet.cell(row=row, column=7, value=rxrate)
        xls_sheet.cell(row=row, column=8, value=ping_output)
        row += 1
    time.sleep(0.3)
    return row

def ping():
    """Collects IP of host for ICMP test, returns formatted Ping command
    """
    ping_host = '1.1.1.1'
    ping_timeout_ms = '350'

    if ping_host == '':
         ping_host = input('\nEnter IP to test [192.0.2.1]: ')
         if ping_host == '': ping_host = '192.0.2.1'
    if operating_system == 'posix': ping_cmd = ('ping -t ' + ping_timeout_ms + ' -c 1 ' + ping_host)
    if operating_system == 'nt': ping_cmd = ('ping -w ' + ping_timeout_ms + ' -n 1 ' + ping_host)
    return(ping_cmd)


if __name__ == '__main__':
    xls = Workbook()
    xls_sheet = xls.active
    xls_filename = 'roam_export.xlsx'
    row = 1
    operating_system = os.name
    ping_command = ping()
    try:
        while True:
            if operating_system == 'posix': row = apple(row)
            if operating_system == 'nt': row = microsoft(row)
    except KeyboardInterrupt:
        try:
            xls.save(xls_filename)
            print('Saved', xls_filename)
        except PermissionError:
            print('File cannot be saved, exiting.')
        print('Script stopped.')
